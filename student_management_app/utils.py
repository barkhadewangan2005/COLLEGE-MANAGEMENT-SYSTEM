"""
Utility functions for the college management system
"""
import logging
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from .models import ActivityLog, Notification
from io.  import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

logger = logging.getLogger(__name__)


def log_activity(user, action, description, ip_address=None):
    """
    Log user activity
    """
    try:
        ActivityLog.objects.create(
            user=user,
            action=action,
            description=description,
            ip_address=ip_address
        )
        logger.info(f"Activity logged: {user.username} - {action}")
    except Exception as e:
        logger.error(f"Error logging activity: {str(e)}")


def create_notification(user, title, message, notification_type='general', link=None):
    """
    Create notification for a user
    """
    try:
        notification = Notification.objects.create(
            user=user,
            title=title,
            message=message,
            notification_type=notification_type,
            link=link
        )
        logger.info(f"Notification created for {user.username}: {title}")
        return notification
    except Exception as e:
        logger.error(f"Error creating notification: {str(e)}")
        return None


def send_email_notification(subject, message, recipient_list, html_message=None):
    """
    Send email notification
    """
    try:
        if html_message:
            email = EmailMultiAlternatives(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=recipient_list
            )
            email.attach_alternative(html_message, "text/html")
            email.send()
        else:
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=recipient_list,
                fail_silently=False,
            )
        logger.info(f"Email sent to {recipient_list}: {subject}")
        return True
    except Exception as e:
        logger.error(f"Error sending email: {str(e)}")
        return False


def send_template_email(subject, template_name, context, recipient_list):
    """
    Send email using template
    """
    try:
        html_message = render_to_string(template_name, context)
        plain_message = strip_tags(html_message)
        
        return send_email_notification(
            subject=subject,
            message=plain_message,
            recipient_list=recipient_list,
            html_message=html_message
        )
    except Exception as e:
        logger.error(f"Error sending template email: {str(e)}")
        return False


def generate_pdf_report(title, data, filename='report.pdf', columns=None):
    """
    Generate PDF report
    
    Args:
        title: Report title
        data: List of dictionaries or list of lists
        filename: Output filename
        columns: List of column headers (optional)
    
    Returns:
        BytesIO buffer with PDF data
    """
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Date
        date_str = datetime.now().strftime("%B %d, %Y %I:%M %p")
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1  # Center
        )
        elements.append(Paragraph(f"Generated: {date_str}", date_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Prepare table data
        if data:
            # Convert dict data to list if needed
            if isinstance(data[0], dict):
                if not columns:
                    columns = list(data[0].keys())
                table_data = [columns]
                for row in data:
                    table_data.append([str(row.get(col, '')) for col in columns])
            else:
                if columns:
                    table_data = [columns] + data
                else:
                    table_data = data
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No data available", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        logger.info(f"PDF report generated: {filename}")
        return buffer
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}")
        return None


def generate_excel_report(title, data, filename='report.xlsx', columns=None):
    """
    Generate Excel report
    
    Args:
        title: Report title
        data: List of dictionaries or list of lists
        filename: Output filename
        columns: List of column headers (optional)
    
    Returns:
        BytesIO buffer with Excel data
    """
    try:
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        
        # Title styling
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color='1a237e')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Date
        ws.merge_cells('A2:D2')
        date_cell = ws['A2']
        date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal='center')
        
        # Prepare data
        start_row = 4
        if data:
            # Convert dict data to list if needed
            if isinstance(data[0], dict):
                if not columns:
                    columns = list(data[0].keys())
                
                # Headers
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=start_row, column=col_num)
                    cell.value = column_title
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Data rows
                for row_num, row_data in enumerate(data, start_row + 1):
                    for col_num, column_title in enumerate(columns, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = row_data.get(column_title, '')
                        cell.alignment = Alignment(horizontal='left')
            else:
                # Headers
                if columns:
                    for col_num, column_title in enumerate(columns, 1):
                        cell = ws.cell(row=start_row, column=col_num)
                        cell.value = column_title
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                    start_row += 1
                
                # Data rows
                for row_num, row_data in enumerate(data, start_row):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.alignment = Alignment(horizontal='left')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Excel report generated: {filename}")
        return buffer
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        return None


def get_client_ip(request):
    """
    Get client IP address from request
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def calculate_percentage(obtained, total):
    """
    Calculate percentage with 2 decimal places
    """
    if total == 0:
        return 0
    return round((obtained / total) * 100, 2)


def format_date_range(start_date, end_date):
    """
    Format date range for display
    """
    try:
        return f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    except Exception as e:
        return f"{start_date} - {end_date}"
